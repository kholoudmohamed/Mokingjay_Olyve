import openpyxl as px


class ExcelDataReader(object):

    # The following function returns a list of columns inside the Excel workbook
    # Append the value of each row inside each column to this list
    @classmethod
    def get_data(cls, fileName, sheetName, HDR=True, numOfHeaders = 1):
        wbook = px.load_workbook(fileName)
        wsheet = wbook.get_sheet_by_name(sheetName)
        if HDR != True:
            StartRow = 1
            numOfHeaders = 0
        else:
            numOfHeaders = numOfHeaders
            StartRow = numOfHeaders + 1
        result = []
        for x in range(1, (wsheet.max_column+1)):
            cols = []
            for y in range(StartRow,(wsheet.max_row+1)):
                cols.append(wsheet.cell(column=x, row=y).value)
            result.append(cols)
        return result